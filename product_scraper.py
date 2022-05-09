import time
import random
from itertools import cycle
from bs4 import BeautifulSoup
from lxml.html import fromstring

import requests
from requests import Session
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

class ProductScraper:

    def __init__(self):
        self.USER_AGENTS = [
            "Mozilla/5.0 (Windows; U; Windows NT 6.1; x64; fr; rv:1.9.2.13) Gecko/20101203 Firebird/3.6.13",
            "Mozilla/5.0 (Windows; U; Windows NT 6.1; rv:2.2) Gecko/20110201",
            "Opera/9.80 (X11; Linux i686; Ubuntu/14.10) Presto/2.12.388 Version/12.16",
            "Mozilla/5.0 (Windows NT 5.2; RW; rv:7.0a1) Gecko/20091211 SeaMonkey/9.23a1pre",
        ]
        self.HEADER = {
            'authority': 'www.amazon.com',
            'pragma': 'no-cache',
            'cache-control': 'no-cache',
            'dnt': '1',
            'upgrade-insecure-requests': '1',
            'user-agent': random.choice(self.USER_AGENTS),
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'none',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-dest': 'document',
            'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
        }



    def get_proxies(self):
        "Get a list of proxies from the free-proxy-list site"
        url = 'https://free-proxy-list.net/'
        response = requests.get(url)
        parser = fromstring(response.text)
        proxies = set()
        for i in parser.xpath('//tbody/tr')[:10]:
            if i.xpath('.//td[7][contains(text(),"yes")]'):
                #Grabbing IP and corresponding PORT
                proxy = ":".join([i.xpath('.//td[1]/text()')[0], i.xpath('.//td[2]/text()')[0]])
                proxies.add(proxy)
        return proxies

    def get_response(self, url, proxy=None):
        """Take a product URL and return the response"""

        # add schema to the url if it doesn't exist
        url = 'https://' + url if not url.startswith('http') else url
        
        # set up the request session
        session = Session()
        header = self.HEADER
        header['user-agent'] = random.choice(self.USER_AGENTS)
        session.headers.update(header)
        print(url)
        # return the response
        try:
            if proxy:
                response = session.get(url, proxies={"http": proxy, "https": proxy})
            else:
                response = session.get(url)

        except requests.exceptions.ConnectionError:
            print("Please Check your Internet Connection!")
            return None
        if response.ok:
            return response


    def scrape_product_info(self, url, proxy=None):
        url = f"https://amazon.com/dp/{url}" if 'amazon' not in url else url
        response = self.get_response(url, proxy=proxy)
        if response:
            soup = BeautifulSoup(response.content, 'html.parser')

            try:
                title = soup.find('span', {'id': 'productTitle'}).text.strip()
            except AttributeError as ae:
                title = None
                print(ae)
            try:
                description = soup.find('div', {'id': 'productDescription'}).p.span.text.strip()
            except AttributeError as ae:
                description = None
                print(ae)

            try:
                details_feature_div = soup.find('div', {'id': 'detailBullets_feature_div'})
                details_list = [' '.join(span.text.split()) for span in details_feature_div.find_all('span', {'class': 'a-list-item'})]
                details = ' | '.join(details_list)
            except AttributeError as ae:
                details = None
                print(ae)

            product_info = {'url': url,'title': title, 'description': description, 'details': details}

            return product_info



class Excel:
    """Work with all excel related tasks."""

    def __init__(self, file_name):
        self.file_name = file_name
        self.font = Font(color="000000", bold=True)
        self.bg_color = PatternFill(fgColor="E8E8E8", fill_type="solid")
        self.customize_excel()

    def create_sheets(self):
        """Create all the sheets required for the project."""
        self.output = (
            self.wb.create_sheet("Output")
            if "Output" not in self.wb.sheetnames
            else self.wb["Output"]
        )

        self.errors = (
            self.wb.create_sheet("Errors")
            if "Errors" not in self.wb.sheetnames
            else self.wb["Errors"]
        )


    def make_columns(self, cells_zip, sheet, general_width=20, url_width=50):
        """Takes zip values of rows and columns and puts values in place with some stylings"""
        # iterating through the column and its values to put them in place
        for col, value in cells_zip:
            cell = sheet[f"{col}1"]
            cell.value = value
            cell.font = self.font
            cell.fill = self.bg_color
            sheet.freeze_panes = cell

            # fixing the column width
            sheet.column_dimensions[col].width = general_width
        # fixing the URL column width
        sheet.column_dimensions["A"].width = url_width

    def customize_output_column(self):
        """Customize the Output column according to its values"""

        # combining columns with its values
        output_column = zip(
            ("A", "B", "C", "D"),
            (
                "Product URL",
                "Product Title",
                "Product Description",
                "Product Details",
            ),
        )
        self.make_columns(output_column, self.output, general_width=50)


    def customize_excel(self):
        """Run all the functions related to excel customization"""
        self.wb = load_workbook(self.file_name)
        self.create_sheets()
        self.customize_output_column()
        self.wb.save(self.file_name)

    def generate_inputs(self):
        """Read the first column of Input sheet and yield the values"""
        inputs = self.wb["Input"]
        for row in range(2, inputs.max_row + 1):
            # generates the links one by one
            if value := inputs[f"A{row}"].value:
                yield value

    def append_output(self, product_info):
        """Read the output and append it to excel"""
        if product_info:
            try:
                self.wb["Output"].append(
                    (
                        product_info["url"],
                        product_info["title"],
                        product_info["description"],
                        product_info["details"],
                    )
                )

                self.wb.save(self.file_name)
            except openpyxl.utils.exceptions.IllegalCharacterError as e:
                self.wb["Errors"].append((product_info["url"], str(e)))



        
    

if __name__ == '__main__':
    filename = 'AmazonProducts.xlsx'
    excel = Excel(filename)
    scraper = ProductScraper()

    
    for url in  excel.generate_inputs():
        print(f"SCRAPING... {url}")
        try:
            product_info = scraper.scrape_product_info(url)
            if product_info:
                excel.append_output(product_info)
            print('sleeping 12 seconds')
            time.sleep(12)
        except Exception as e:
            excel.wb["Errors"].append((url, str(e)))
            print(e)
    excel.wb.save(filename)
    
    